import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import re

def json_to_excel(input_file, output_file):
    def escape_excel_formula(val):
        if isinstance(val, str):
            val = val.strip()
            if val.startswith("=") or val.startswith("=>") or val.startswith("=<") or val.startswith("<=") or val.startswith(">="):
                return "'" + val
        return val

    def normalize_number(value):
        try:
            return float(str(value).replace(",", "."))
        except:
            return value

    def clean_excel_string(s):
        if not isinstance(s, str):
            return escape_excel_formula(s)
        s = s.replace("\r", "")
        s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
        return escape_excel_formula(s)

    with open(input_file, 'r', encoding='utf-8') as f:
        json_data = json.load(f)

    stoffliste = json_data["gefaehrlicheInhaltsstoffe"]
    index_to_name = {i: s["stoffname"] for i, s in enumerate(stoffliste)}
    name_to_substanz = {s["stoffname"]: s for s in stoffliste}

    gefaehrlicheInhaltsstoffe = []
    for s in stoffliste:
        nummern_kombi = f"CAS: {s['casNummer']}\nEG: {s['egNummer']}\nIndex: {s['indexNummer']}\nREACH: {s['reachNummer']}"

        gefahr_hsaetze = []
        gefahr_kategorien = s.get("gefahrenkategorien", [])
        h_saetze = s.get("hSaetze", [])

        for i in range(min(len(gefahr_kategorien), len(h_saetze))):
            gefahr_hsaetze.append(f"{gefahr_kategorien[i]};{h_saetze[i]}")
        for i in range(len(h_saetze), len(gefahr_kategorien)):
            gefahr_hsaetze.append(f"{gefahr_kategorien[i]};")
        for i in range(len(gefahr_kategorien), len(h_saetze)):
            gefahr_hsaetze.append(f";{h_saetze[i]}")

        gefahr_hsaetze_final = []
        for item in gefahr_hsaetze:
            if item:
                parts = item.split(";")
                gefahr_kategorie = parts[0] if len(parts) > 0 else ""
                h_satz = parts[1] if len(parts) > 1 else ""
                gefahr_hsaetze_final.append(f"{h_satz}-{gefahr_kategorie}")
            else:
                gefahr_hsaetze_final.append("")

        substance = {
            "Stoffname": clean_excel_string(s.get("stoffname", "")),
            "Nummern": clean_excel_string(nummern_kombi) if nummern_kombi else "",
            "Konzentration": escape_excel_formula(s.get("konzentration", "")),
            "Konzentrationseinheit": clean_excel_string(s.get("konzentrationEinheit", "")),
            "EUH-Sätze": clean_excel_string(";\n".join(s.get("euhSaetze", []))) if s.get("euhSaetze") else "",
            "Gefahrenkategorien und H-Sätze": clean_excel_string(";\n".join(gefahr_hsaetze_final)) if gefahr_hsaetze_final else "",
            "M-Faktor Akut": clean_excel_string(s.get("mFaktorAkut", "")),
            "M-Faktor Chronisch": clean_excel_string(s.get("mFaktorChronisch", ""))
        }

        if "spezKonzGrenzen" in s:
            grenzen = []
            for grenze in s["spezKonzGrenzen"]:
                grenzen.append(f"{grenze['hSaetze']}-{grenze['gefahrenkategorie']}: {grenze['spezKonzGrenze']}")
            substance["Spezifische Konzentrationsgrenzen"] = clean_excel_string(";\n".join(grenzen))
        else:
            substance["Spezifische Konzentrationsgrenzen"] = ""

        gefaehrlicheInhaltsstoffe.append(substance)

    tox = []
    for t in json_data.get("tox", []):
        index = t["substanceIndex"]
        stoffname = index_to_name.get(index, "Unbekannt")
        tox.append({
            "Substanz": clean_excel_string(f"{stoffname} ({index})"),
            "Typ": clean_excel_string(t["typ"]),
            "Expositionsweg": clean_excel_string(t["expositionsweg"]),
            "Wert": normalize_number(t["wert"]),
            "Einheit": clean_excel_string(t["einheit"]),
            "Methode": clean_excel_string(t["methode"]),
            "Spezies": clean_excel_string(t["spezies"]),
            "Expositionsdauer": clean_excel_string(t["expositionsdauer"]),
        })

    ecoTox = []
    for e in json_data.get("ecoTox", []):
        index = e["substanceIndex"]
        stoffname = index_to_name.get(index, "Unbekannt")
        ecoTox.append({
            "Substanz": clean_excel_string(f"{stoffname} ({index})"),
            "Typ": clean_excel_string(e["typ"]),
            "Spezies": clean_excel_string(e["spezies"]),
            "Wert": normalize_number(e["wert"]),
            "Einheit": clean_excel_string(e["einheit"]),
            "Methode": clean_excel_string(e["methode"]),
            "Trophieebene": clean_excel_string(e["trophieebene"]),
            "Expositionsdauer": clean_excel_string(e["expositionsdauer"]),
        })

    bcf_logkow_abbau = []
    for s in stoffliste:
        bcf_logkow_abbau.append({
            "Stoffname": clean_excel_string(s["stoffname"]),
            "BCF": clean_excel_string(s.get("bcf", "")),
            "LogKow": clean_excel_string(s.get("logKow", "")),
            "Biologisch abbaubar": clean_excel_string(s.get("biolAbbaubar", ""))
        })

    # Export nach Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df1 = pd.DataFrame(gefaehrlicheInhaltsstoffe)
        df2 = pd.DataFrame(tox)
        df3 = pd.DataFrame(ecoTox)
        df4 = pd.DataFrame(bcf_logkow_abbau)

        df1.to_excel(writer, sheet_name="Daten", index=False, startrow=0)
        df2.to_excel(writer, sheet_name="Daten", index=False, startrow=len(df1) + 2)
        df3.to_excel(writer, sheet_name="Daten", index=False, startrow=len(df1) + len(df2) + 4)
        df4.to_excel(writer, sheet_name="Daten", index=False, startrow=len(df1) + len(df2) + len(df3) + 6)

        workbook = writer.book
        worksheet = workbook['Daten']

        column_widths = {
            'A': 19, 'B': 23, 'C': 15, 'D': 8, 'E': 8,
            'F': 25, 'G': 11, 'H': 8, 'I': 25
        }

        for column, width in column_widths.items():
            worksheet.column_dimensions[column].width = width

        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        for row_index in worksheet.iter_rows():
            worksheet.row_dimensions[row_index[0].row].height = None
input_file =r"C:\Users\Steffen Kades\Desktop\Blatt11\final.json"
output_file =r"C:\Users\Steffen Kades\Desktop\Blatt11\final.xlsx"
json_to_excel(input_file, output_file)