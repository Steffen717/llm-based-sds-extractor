import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import fitz  # PyMuPDF
import threading
import time
from PIL import Image, ImageTk
import os
from openpyxl.styles import PatternFill
from excel_to_JSON import excel_to_json

EXCEL_ROW_HEIGHT_PX = 15
MIN_ROWS = 2

class ExcelViewer:
    def __init__(self, root):
        self.root = root
        self.entries = {}
        self.current_page = 0
        self.current_dataset_index = 0
        self.dataset_folders = []
        self.changes = 0
        self.total_errors = 0
        self.zoom_level = 0.925
        self.accumulated_time = 0      
        self.last_start_time = None     
        self.timer_running = False      
        self.dataset_errors = {}
        self.dataset_times = {}
        self.dataset_total_cells = {}
        self.dataset_row_count = {}
        self.setup_gui()
        self.base_folder = filedialog.askdirectory(title="Wähle den Hauptordner mit Datensätzen")
        if self.base_folder:
            self.dataset_folders = sorted([
                os.path.join(self.base_folder, d)
                for d in os.listdir(self.base_folder)
                if os.path.isdir(os.path.join(self.base_folder, d))
            ])
            if self.dataset_folders:
                self.load_dataset(0)
                self.start_timer()

    def setup_gui(self):
        self.root.title("Evaluation")
        self.root.geometry("1400x900")
        self.root.bind("<F11>", self.toggle_fullscreen)
        self.root.bind("<Configure>", self.resize_frames)
        self.left_frame = tk.Frame(self.root)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.right_frame = tk.Frame(self.root)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Excel-Bereich
        self.canvas = tk.Canvas(self.left_frame)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.scrollbar = tk.Scrollbar(self.left_frame, orient=tk.VERTICAL, command=self.canvas.yview)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        self.table_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.table_frame, anchor='nw')
        self.table_frame.bind("<Configure>", lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all")))
        self.canvas.bind("<Enter>", self.enable_excel_scroll)
        self.canvas.bind("<Leave>", self.disable_excel_scroll)

        # Steuerung und Fehleranzeige
        self.control_frame = tk.Frame(self.right_frame)
        self.control_frame.pack(fill=tk.X, pady=5)
        self.control_inner = tk.Frame(self.control_frame)
        self.control_inner.pack(anchor="center")
        self.error_label = tk.Label(self.control_inner, text="Fehlermenge: 0", font=('Arial', 12))
        self.error_label.grid(row=0, column=1, padx=5, columnspan=2)
        self.total_error_label = tk.Label(self.control_inner, text="Gesamtfehler: 0", font=('Arial', 12))
        self.total_error_label.grid(row=0, column=3, padx=5, columnspan=2)
        self.remove_error_button = tk.Button(self.control_inner, text="Fehler abziehen", command=self.remove_error)
        self.remove_error_button.grid(row=0, column=5, padx=5)
        self.add_error_button = tk.Button(self.control_inner, text="Fehler hinzufügen", command=self.add_error)
        self.add_error_button.grid(row=0, column=6, padx=5)

        # Navigation und Timer
        self.nav_frame = tk.Frame(self.right_frame)
        self.nav_frame.pack(pady=10)
        self.nav_inner = tk.Frame(self.nav_frame)
        self.nav_inner.pack(anchor="center")
        self.search_frame = tk.Frame(self.right_frame)
        self.search_frame.pack(pady=5)
        self.search_label = tk.Label(self.search_frame, text="PDF durchsuchen:")
        self.search_label.pack(side=tk.LEFT)
        self.search_entry = tk.Entry(self.search_frame, width=20)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_button = tk.Button(self.search_frame, text="Suchen", command=self.search_pdf)
        self.search_button.pack(side=tk.LEFT)
        self.next_hit_button = tk.Button(self.search_frame, text="Weiter", command=self.next_hit, state=tk.DISABLED)
        self.next_hit_button.pack(side=tk.LEFT)
        self.prev_data_button = tk.Button(self.nav_inner, text="◀ Zurück", command=self.load_prev_dataset)
        self.prev_data_button.grid(row=0, column=0, padx=5)
        self.current_dataset_label = tk.Label(self.nav_inner, text="0 von 0", font=('Arial', 10))
        self.current_dataset_label.grid(row=0, column=1, padx=5)
        self.next_data_button = tk.Button(self.nav_inner, text="Weiter ▶", command=lambda: (self.save_excel(), self.load_next_dataset()))
        self.next_data_button.grid(row=0, column=2, padx=5)
        self.save_button = tk.Button(self.nav_inner, text="Änderungen speichern", command=self.save_excel)
        self.save_button.grid(row=0, column=3, padx=5)
        self.export_button = tk.Button(self.nav_inner, text="Auswertung exportieren", command=lambda:((self.save_excel(), self.save_dataset_results())))
        self.export_button.grid(row=0, column=7, padx=5)
        self.timer_label = tk.Label(self.nav_inner, text="Zeit: 0s", font=('Arial', 10))
        self.timer_label.grid(row=0, column=4, padx=5)
        self.pause_button = tk.Button(self.nav_inner, text="Stopp", command=self.pause_timer)
        self.pause_button.grid(row=0, column=5, padx=2)
        self.resume_button = tk.Button(self.nav_inner, text="Fortsetzen", command=self.resume_timer, state=tk.DISABLED)
        self.resume_button.grid(row=0, column=6, padx=2)

        # PDF-Bereich
        self.pdf_frame = tk.Frame(self.right_frame)
        self.pdf_frame.pack(fill=tk.BOTH, expand=True)
        self.pdf_canvas = tk.Canvas(self.pdf_frame, bg="white")
        self.pdf_canvas.grid(row=0, column=0, sticky="nsew")
        self.pdf_v_scroll = tk.Scrollbar(self.pdf_frame, orient=tk.VERTICAL, command=self.pdf_canvas.yview)
        self.pdf_v_scroll.grid(row=0, column=1, sticky="ns")
        self.pdf_h_scroll = tk.Scrollbar(self.pdf_frame, orient=tk.HORIZONTAL, command=self.pdf_canvas.xview)
        self.pdf_h_scroll.grid(row=1, column=0, sticky="ew")
        self.pdf_canvas.configure(
            yscrollcommand=self.pdf_v_scroll.set,
            xscrollcommand=self.pdf_h_scroll.set
        )
        self.pdf_frame.grid_rowconfigure(0, weight=1)
        self.pdf_frame.grid_columnconfigure(0, weight=1)
        self.pdf_canvas.bind("<Enter>", self.enable_pdf_scroll)
        self.pdf_canvas.bind("<Leave>", self.disable_scroll_focus)

    def enable_excel_scroll(self, event=None):
        self.canvas.bind_all("<MouseWheel>", self.scroll_excel)
    def disable_excel_scroll(self, event=None):
        self.canvas.unbind_all("<MouseWheel>")

    def pause_timer(self):
        if not self.timer_running:
            return
        now = time.time()
        self.accumulated_time += now - self.last_start_time
        self.timer_running = False
        self.pause_button.config(state=tk.DISABLED)
        self.resume_button.config(state=tk.NORMAL)

    def resume_timer(self):
        if self.timer_running:
            return
        self.last_start_time = time.time()
        self.timer_running = True
        self.pause_button.config(state=tk.NORMAL)
        self.resume_button.config(state=tk.DISABLED)

    def toggle_fullscreen(self, event=None):
        self.fullscreen = not getattr(self, 'fullscreen', False)
        self.root.attributes("-fullscreen", self.fullscreen)

    def resize_frames(self, event=None):
        total_width = self.root.winfo_width()
        left_width = int(total_width * 0.575)
        right_width = total_width - left_width
        self.left_frame.config(width=left_width)
        self.right_frame.config(width=right_width)

    def scroll_excel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def enable_pdf_scroll(self, event=None):
        self.pdf_canvas.bind_all("<Control-MouseWheel>", self.zoom_with_mouse)
        self.pdf_canvas.bind_all("<MouseWheel>", self.scroll_pdf)
        self.pdf_canvas.bind_all("<Shift-MouseWheel>", self.scroll_pdf_x)

    def scroll_pdf_x(self, event):
        self.pdf_canvas.xview_scroll(int(-1 * (event.delta / 120)), "units")
    def disable_scroll_focus(self, event=None):
        self.pdf_canvas.unbind_all("<MouseWheel>")
        self.pdf_canvas.unbind_all("<Shift-MouseWheel>")
        self.pdf_canvas.unbind_all("<Control-MouseWheel>")
    def zoom_with_mouse(self, event):
        if event.delta > 0:
            self.zoom_in()
        else:
            self.zoom_out()
    def scroll_pdf(self, event):
        self.pdf_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def zoom_in(self):
        self.zoom_level += 0.1
        self.render_pdf()
    def zoom_out(self):
        if self.zoom_level > 0.2:
            self.zoom_level -= 0.1
            self.render_pdf()

    def count_rows(self, worksheet):
        table_widths = [9, 8, 8, 4]
        header_row_counts = [1, 1, 1, 1]
        total_rows = 0
        table_idx = 0
        in_table = False
        skipped_headers = 0
        empty_streak = 0

        for row in worksheet.iter_rows():
            if table_idx >= len(table_widths):
                break

            width = table_widths[table_idx]
            is_empty = all(cell.value is None for cell in row[:width])

            if not in_table:
                if is_empty:
                    empty_streak += 1
                    if empty_streak == 2:
                        table_idx += 1
                        empty_streak = 0
                else:
                    in_table = True
                    skipped_headers = 0
                    empty_streak = 0
                    total_rows += 1
            else:
                if skipped_headers < header_row_counts[table_idx]:
                    skipped_headers += 1
                else:
                    if is_empty:
                        table_idx += 1
                        in_table = False
                        empty_streak = 1
                    else:
                        total_rows += 1

        return total_rows

    def count_cells(self, worksheet):
        table_widths = [9, 8, 8, 4]
        header_row_counts = [1, 1, 1, 1]
        total_cells = 0
        table_idx = 0
        in_table = False
        skipped_headers = 0
        data_rows = 0
        empty_streak = 0

        for row in worksheet.iter_rows():
            if table_idx >= len(table_widths):
                break

            width = table_widths[table_idx]
            is_empty = all(cell.value is None for cell in row[:width])

            if not in_table:
                if is_empty:
                    empty_streak += 1
                    if empty_streak == 2:
                        table_idx += 1
                        empty_streak = 0
                else:
                    in_table = True
                    skipped_headers = 0
                    data_rows = 0
                    empty_streak = 0
                    total_cells += width
            else:
                if skipped_headers < header_row_counts[table_idx]:
                    skipped_headers += 1
                else:
                    if is_empty:
                        total_cells += data_rows * width
                        table_idx += 1
                        in_table = False
                        empty_streak = 1
                    else:
                        data_rows += 1

        if in_table and table_idx < len(table_widths):
            total_cells += data_rows * table_widths[table_idx]

        return total_cells

    def load_excel(self, path):
        wb = load_workbook(path)
        ws = wb.active
        self.current_total_cells = self.count_cells(ws)
        self.current_row_count = self.count_rows(ws)
        self.dataset_row_count[os.path.basename(self.current_folder)] = self.current_row_count

        for widget in self.table_frame.winfo_children():
            widget.destroy()
        self.entries.clear()
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            row_h = ws.row_dimensions[r_idx].height or MIN_ROWS * EXCEL_ROW_HEIGHT_PX
            if row_h < MIN_ROWS * EXCEL_ROW_HEIGHT_PX:
                row_h = MIN_ROWS * EXCEL_ROW_HEIGHT_PX
            for c_idx, cell in enumerate(row, start=1):
                col_l = cell.column_letter
                col_w = (ws.column_dimensions[col_l].width or 8) * 7
                entry = tk.Text(self.table_frame, wrap=tk.WORD,
                                height=int(row_h / EXCEL_ROW_HEIGHT_PX),
                                width=int(col_w / 7), font=('Arial', 10))
                entry.insert("1.0", str(cell.value) if cell.value is not None else "")
                entry.original_text = entry.get("1.0", "end-1c")
                entry.grid(row=r_idx, column=c_idx, sticky="nsew", padx=1, pady=1)
                entry.bind("<FocusOut>", self.check_changes)
                self.entries[(r_idx, c_idx)] = entry
        self.table_frame.update_idletasks()

    def load_pdf(self, path):
        self.pdf_path = path
        self.render_pdf()

    def search_pdf(self):
        search_term = self.search_entry.get().strip()
        if not search_term or not hasattr(self, 'doc'):
            return
        self.search_hits = []
        for page_num, page in enumerate(self.doc):
            areas = page.search_for(search_term)
            for rect in areas:
                self.search_hits.append((page_num, rect))
        if not self.search_hits:
            messagebox.showinfo("Keine Treffer", f"'{search_term}' wurde nicht gefunden.")
            self.next_hit_button.config(state=tk.DISABLED)
            return
        self.current_hit_index = 0
        self.next_hit_button.config(state=tk.NORMAL)
        self.render_pdf(highlights=self.search_hits, scroll_to_hit=self.search_hits[0])

    def next_hit(self):
        if not hasattr(self, 'search_hits') or not self.search_hits:
            return
        self.current_hit_index = (self.current_hit_index + 1) % len(self.search_hits)
        next_hit = self.search_hits[self.current_hit_index]
        self.render_pdf(highlights=self.search_hits, scroll_to_hit=next_hit)

    def render_pdf(self, highlights=None, scroll_to_hit=None):
        self.doc = fitz.open(self.pdf_path)
        self.pdf_canvas.delete("all")
        self.pdf_images = []
        y_offset = 0
        self.page_positions = []

        for page_index, page in enumerate(self.doc):
            mat = fitz.Matrix(1.5 * self.zoom_level, 1.5 * self.zoom_level)
            pix = page.get_pixmap(matrix=mat)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            photo = ImageTk.PhotoImage(img)
            self.pdf_canvas.create_image(0, y_offset, image=photo, anchor=tk.NW)
            self.pdf_images.append(photo)
            self.page_positions.append(y_offset)

            if highlights:
                for pnum, rect in highlights:
                    if pnum == page_index:
                        rect = rect * mat
                        x0, y0, x1, y1 = rect.x0, rect.y0, rect.x1, rect.y1
                        self.pdf_canvas.create_rectangle(x0, y_offset + y0, x1, y_offset + y1, outline="red", width=2)

            y_offset += img.height + 10

        self.pdf_canvas.config(scrollregion=self.pdf_canvas.bbox("all"))

        if scroll_to_hit:
            pnum, rect = scroll_to_hit
            mat = fitz.Matrix(1.5 * self.zoom_level, 1.5 * self.zoom_level)
            rect = rect * mat
            scroll_y = self.page_positions[pnum] + rect.y0
            self.pdf_canvas.yview_moveto(scroll_y / (y_offset + 1))

    def check_changes(self, event):
        for key, entry in self.entries.items():
            curr = entry.get("1.0", "end-1c")
            if curr != getattr(entry, "original_text", "") and not getattr(entry, "changed", False):
                entry.config(bg="lightyellow")
                entry.changed = True
                self.changes += 1
                folder = os.path.basename(self.current_folder)
                self.dataset_errors[folder] = self.dataset_errors.get(folder, 0) + 1
                self.update_error_labels()

    def update_error_labels(self):
        self.total_errors = sum(self.dataset_errors.values())
        self.error_label.config(text=f"Fehlermenge: {self.changes}")
        self.total_error_label.config(text=f"Gesamtfehler: {self.total_errors}")

    def remove_error(self):
        if self.changes > 0:
            self.changes -= 1
        folder = os.path.basename(self.current_folder)
        if self.dataset_errors.get(folder, 0) > 0:
            self.dataset_errors[folder] -= 1
        self.update_error_labels()

    def add_error(self):
        self.changes += 1
        folder = os.path.basename(self.current_folder)
        self.dataset_errors[folder] = self.dataset_errors.get(folder, 0) + 1
        self.update_error_labels()

    def start_timer(self):
        # Einmalig Thread starten
        self.last_start_time = time.time()
        self.timer_running = True
        def update_time():
            while True:
                if self.timer_running:
                    now = time.time()
                    elapsed = int(self.accumulated_time + (now - self.last_start_time))
                else:
                    elapsed = int(self.accumulated_time)
                self.timer_label.config(text=f"Zeit: {elapsed}s")
                if hasattr(self, "current_folder"):
                    fn = os.path.basename(self.current_folder)
                    self.dataset_times[fn] = elapsed
                time.sleep(1)
        threading.Thread(target=update_time, daemon=True).start()

    def load_dataset(self, index):
        self.current_dataset_index = index
        folder = self.dataset_folders[index]
        self.current_folder = folder
        excel_path = os.path.join(folder, "tabelle.xlsx")
        pdf_path = os.path.join(folder, "dokument.pdf")
        if not os.path.exists(excel_path) or not os.path.exists(pdf_path):
            messagebox.showerror("Fehler", f"Fehlende Dateien in {folder}")
            return

        # Reset für neues Dataset
        self.changes = 0
        self.error_label.config(text="Fehlermenge: 0")

        # Timer zurücksetzen
        self.accumulated_time = 0
        self.last_start_time = time.time()
        self.timer_running = True
        self.pause_button.config(state=tk.NORMAL)
        self.resume_button.config(state=tk.DISABLED)

        # Excel und PDF laden
        self.load_excel(excel_path)
        self.load_pdf(pdf_path)
        first = self.entries.get((1, 1))
        if first:
            self.root.after(50, lambda: (self.root.focus_force(), first.focus_set()))
        # Zellenanzahl speichern
        self.dataset_total_cells[os.path.basename(folder)] = self.current_total_cells
        self.update_current_dataset_label()

    def update_current_dataset_label(self):
        self.current_dataset_label.config(text=f"{self.current_dataset_index + 1} von {len(self.dataset_folders)}")

    def load_next_dataset(self):
        self.resume_timer
        self.save_excel
        if self.current_dataset_index + 1 < len(self.dataset_folders):
            self.load_dataset(self.current_dataset_index + 1)
        else:
            messagebox.showinfo("Ende", "Keine weiteren Datensätze vorhanden.")

    def load_prev_dataset(self):
        if self.current_dataset_index > 0:
            self.load_dataset(self.current_dataset_index - 1)
        else:
            messagebox.showinfo("Anfang", "Du bist beim ersten Datensatz.")

    def save_excel(self):
        orig_path = os.path.join(self.current_folder, "tabelle.xlsx")
        wb = load_workbook(orig_path)
        ws = wb.active
        fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
        for (r, c), entry in self.entries.items():
            value = entry.get("1.0", "end-1c")
            cell = ws.cell(row=r, column=c, value=value)
            if getattr(entry, "changed", False):
                cell.fill = fill
        save_path = os.path.join(self.current_folder, "tabelle_geaendert.xlsx")
        wb.save(save_path)
        json_filename = "corrected_final.json"
        json_path = os.path.join(self.current_folder, json_filename)
        excel_to_json(orig_path, json_path)

    def save_dataset_results(self):
        
        result_file = os.path.join(self.base_folder, "auswertung.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Fehler", "Zeit (s)", "Zellen", "Zeilen"])
        for folder in self.dataset_folders:
            name = os.path.basename(folder)
            fehler = self.dataset_errors.get(name, 0)
            zeit = self.dataset_times.get(name, 0)
            zellen = self.dataset_total_cells.get(name, 0)
            zeilen = self.dataset_row_count.get(name, 0)
            ws.append([name, fehler, zeit, zellen, zeilen])
        wb.save(result_file)
        messagebox.showinfo("Auswertung gespeichert", f"Ergebnisse wurden in {result_file} gespeichert.")

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelViewer(root)
    root.mainloop()
